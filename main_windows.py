import serial
import csv
import serial.tools.list_ports
from dbfread import DBF
from openpyxl import Workbook
from datetime import datetime
import openpyxl
import os
import time

def read_serial_and_write_csv():
    # Muestra los puertos seriales disponibles
    puertos_disponibles = serial.tools.list_ports.comports()
    for puerto in puertos_disponibles:
        print(puerto.device)

    print('-------------------------------------------')
    puerto = 'COM3'  # Reemplaza 'COM3' con el nombre de tu puerto serial

    baudios = 9600  # Ajusta la velocidad de baudios según tu configuración
    archivo_csv = './peso.csv'  # Nombre del archivo CSV a crear/sobrescribir

    try:
        ser = serial.Serial(puerto, baudios)
        print('Conexión establecida en el puerto', puerto)

        while True:
            if ser.in_waiting > 0:
                datos = ser.readline().decode().strip()
                print('Datos recibidos:', datos)

                with open(archivo_csv, 'w', newline='') as archivo:
                    writer = csv.writer(archivo)
                    writer.writerow([datos])

    except serial.SerialException as e:
        print('Error de conexión:', str(e))


def convert_databases_to_excel():
    while True:
        # Rutas de los archivos .dbf
        dbf_paths = [
            '../bd/ipedidoc.dbf',
            '../bd/ipedidod.dbf',
            '../bd/oprod.dbf',
            '../bd/ordproc.dbf',
            '../bd/PEDIENTR.dbf',
            '../bd/remc.dbf',
            '../bd/remd.dbf',
        ]

        for i, dbf_path in enumerate(dbf_paths, start=1):
            base_name = os.path.splitext(os.path.basename(dbf_path))[0]

            # Abre el archivo .dbf
            dbf = DBF(dbf_path, ignore_missing_memofile=True)

            # Crea una lista para almacenar los datos
            datos = []

            # Lee los registros del archivo .dbf
            for record in dbf:
                datos.append(record)

            # Crea un nuevo archivo de Excel
            wb = Workbook()

            # Selecciona la hoja activa
            sheet = wb.active

            # Escribe los encabezados de las columnas en la primera fila
            encabezados = dbf.field_names
            for col, encabezado in enumerate(encabezados, start=1):
                sheet.cell(row=1, column=col, value=encabezado)

            # Escribe los datos en las filas siguientes
            for row, record in enumerate(datos, start=2):
                for col, campo in enumerate(encabezados, start=1):
                    sheet.cell(row=row, column=col, value=record[campo])

            # Guarda el archivo de Excel
            excel_path = f'./excel_files/{base_name}.xlsx'
            wb.save(excel_path)

        excel_folder = './excel_files/'

        # Obtén la lista de archivos de Excel en la carpeta
        excel_files = [f for f in os.listdir(excel_folder) if f.endswith('.xlsx')]

        # Crea un nuevo archivo de Excel
        merged_wb = Workbook()

        # Itera sobre los archivos de Excel y agrega las hojas al archivo fusionado
        for file in excel_files:
            file_path = os.path.join(excel_folder, file)
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                new_sheet = merged_wb.create_sheet(title=f"{file[:-5]} - {sheet_name}")  # Utiliza el nombre del archivo como parte del título de la hoja
                for row in sheet.iter_rows(values_only=True):
                    new_sheet.append(row)

        # Elimina la hoja de inicio vacía
        del merged_wb['Sheet']

        # Obtener la fecha y hora actual
        fecha_hora_actual = datetime.now()

        # Formatear la fecha y hora como un string
        fecha_hora_str = fecha_hora_actual.strftime("%Y-%m-%d %H:%M:%S")

        # Guarda el archivo fusionado con la fecha y hora al momento de generación
        merged_file_path = f"./merged.xlsx"
        merged_wb.save(merged_file_path)

        # Wait for 10 seconds before generating the merged file again
        time.sleep(1)


# Ejecuta las funciones en hilos separados
import threading

# Función para leer el puerto serial y escribir en CSV
serial_thread = threading.Thread(target=read_serial_and_write_csv)

# Función para convertir bases de datos a Excel y generar el archivo fusionado cada 10 segundos
excel_thread = threading.Thread(target=convert_databases_to_excel)

# Inicia ambos hilos
serial_thread.start()
excel_thread.start()