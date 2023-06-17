from dbfread import DBF
from openpyxl import Workbook
from datetime import datetime
import openpyxl
import os


def convert_databases_to_excel():
  # Rutas de los archivos .dbf
  dbf_paths = [
  '/Users/brito/Documents/datostablas/ipedidoc.dbf',
  '/Users/brito/Documents/datostablas/ipedidod.dbf',
  '/Users/brito/Documents/datostablas/oprod.dbf',
  '/Users/brito/Documents/datostablas/ordproc.dbf',
  '/Users/brito/Documents/datostablas/PEDIENTR.dbf',
  '/Users/brito/Documents/datostablas/remc.dbf',
  '/Users/brito/Documents/datostablas/remd.dbf',
  ]

  for i, dbf_path in enumerate(dbf_paths, start=1):

        base_name = os.path.splitext(os.path.basename(dbf_path))[0]

        # Abre el archivo .dbf
        dbf = DBF(dbf_path, ignore_missing_memofile=True)

        # Crea una lista para almacenar los datos
        datos = []

        # Lee los registros del archivo .dbf
        for record in dbf:
            datos.append(record)

        # Crea un nuevo archivo de Excel
        wb = Workbook()

        # Selecciona la hoja activa
        sheet = wb.active

        # Escribe los encabezados de las columnas en la primera fila
        encabezados = dbf.field_names
        for col, encabezado in enumerate(encabezados, start=1):
            sheet.cell(row=1, column=col, value=encabezado)

        # Escribe los datos en las filas siguientes
        for row, record in enumerate(datos, start=2):
            for col, campo in enumerate(encabezados, start=1):
                sheet.cell(row=row, column=col, value=record[campo])

        # Guarda el archivo de Excel
        excel_path = f'/Users/brito/excel_creator/excel_files/{base_name}.xlsx'
        wb.save(excel_path)


def merge_files():
    excel_folder = '/Users/brito/excel_creator/excel_files/'

    # Obtén la lista de archivos de Excel en la carpeta
    excel_files = [f for f in os.listdir(excel_folder) if f.endswith('.xlsx')]

    # Crea un nuevo archivo de Excel
    merged_wb = Workbook()

    # Itera sobre los archivos de Excel y agrega las hojas al archivo fusionado
    for file in excel_files:
        file_path = os.path.join(excel_folder, file)
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            new_sheet = merged_wb.create_sheet(title=f"{file[:-5]} - {sheet_name}")  # Utiliza el nombre del archivo como parte del título de la hoja
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)

    # Elimina la hoja de inicio vacía
    del merged_wb['Sheet']

    # Obtener la fecha y hora actual
    fecha_hora_actual = datetime.now()

    # Formatear la fecha y hora como un string
    fecha_hora_str = fecha_hora_actual.strftime("%Y-%m-%d %H:%M:%S")

    # Guarda el archivo fusionado con la fecha y hora al momento de generación 
    # merged_file_path = f"/Users/brito/excel_creator/{fecha_hora_str}.xlsx"

    merged_file_path = f"/Users/brito/excel_creator/merged.xlsx"
    merged_wb.save(merged_file_path)

convert_databases_to_excel()
merge_files()